import io
from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import curve_fit, minimize


st.set_page_config(page_title="Paid Search Mix Optimizer", layout="wide")


REQUIRED_FIELDS = [
    ("date", "Date"),
    ("campaign_key", "Campaign ID or Campaign Name"),
    ("campaign_name", "Campaign Name"),
    ("cost", "Spend / Cost"),
    ("conversions", "Conversions"),
]

OPTIONAL_FIELDS = [
    ("clicks", "Clicks"),
    ("impressions", "Impressions"),
    ("impression_share", "Impression Share"),
    ("conversion_value", "Conversion Value / Revenue"),
    ("category", "Category (if already mapped)"),
]

SYNONYM_MAP = {
    "date": ["date", "day"],
    "campaign_key": ["campaign_key", "campaign id", "campaign_id", "campaign", "campaign name", "campaign_name"],
    "campaign_name": ["campaign_name", "campaign name", "campaign"],
    "cost": ["cost", "spend", "amount spent"],
    "conversions": ["conversions", "conv", "orders", "leads"],
    "clicks": ["clicks"],
    "impressions": ["impressions", "impr"],
    "impression_share": ["impression share", "impression_share", "search impr. share", "search impression share"],
    "conversion_value": ["conversion_value", "conversion value", "revenue", "sales"],
    "category": ["category", "bucket", "channel group"],
}


@dataclass
class CampaignModel:
    campaign_key: str
    campaign_name: str
    category: str
    a: float
    b: float
    fit_status: str
    data_points: int
    recent_avg_spend: float
    recent_avg_conversions: float

    def predict(self, spend: float, factor: float = 1.0) -> float:
        spend = max(float(spend), 0.0)
        return max(float(hill_curve(spend, self.a, self.b) * factor), 0.0)

    def marginal(self, spend: float, factor: float = 1.0) -> float:
        spend = max(float(spend), 0.0)
        return max(float(hill_curve_derivative(spend, self.a, self.b) * factor), 0.0)



def hill_curve(spend, a, b):
    spend = np.asarray(spend, dtype=float)
    return a * spend / (b + spend + 1e-9)



def hill_curve_derivative(spend, a, b):
    spend = np.asarray(spend, dtype=float)
    return a * b / np.maximum((b + spend + 1e-9) ** 2, 1e-9)



def read_uploaded_csv(uploaded_file):
    raw_bytes = uploaded_file.getvalue()
    return pd.read_csv(io.BytesIO(raw_bytes))



def normalize_impression_share(series):
    if series is None:
        return pd.Series(dtype=float)
    s = pd.to_numeric(series, errors="coerce")
    if s.dropna().empty:
        return s
    if s.max(skipna=True) > 1.5:
        s = s / 100.0
    return s.clip(lower=0.0, upper=1.0)



def to_numeric(series):
    return pd.to_numeric(series, errors="coerce")



def monday_of_week(date_series):
    return date_series - pd.to_timedelta(date_series.dt.weekday, unit="D")



def clean_column_name(name: str) -> str:
    return str(name).strip().lower().replace("_", " ")



def guess_column(field_key: str, columns: List[str], required: bool) -> str:
    cleaned_cols = {col: clean_column_name(col) for col in columns}
    candidates = SYNONYM_MAP.get(field_key, [field_key])
    for col, cleaned in cleaned_cols.items():
        if cleaned in candidates:
            return col
    for col, cleaned in cleaned_cols.items():
        if any(candidate in cleaned for candidate in candidates):
            return col
    if field_key == "campaign_name" and "campaign_key" in columns:
        return columns[0]
    return columns[0] if required else "(not used)"



def standardize_dataset(df, mapping):
    clean = pd.DataFrame()

    clean["date"] = pd.to_datetime(df[mapping["date"]], errors="coerce")
    clean["campaign_key"] = df[mapping["campaign_key"]].astype(str).str.strip()
    clean["campaign_name"] = df[mapping["campaign_name"]].astype(str).str.strip()
    clean["cost"] = to_numeric(df[mapping["cost"]]).fillna(0.0)
    clean["conversions"] = to_numeric(df[mapping["conversions"]]).fillna(0.0)

    for field in ["clicks", "impressions", "conversion_value"]:
        source = mapping.get(field)
        if source and source != "(not used)":
            clean[field] = to_numeric(df[source]).fillna(0.0)
        else:
            clean[field] = 0.0

    is_col = mapping.get("impression_share")
    if is_col and is_col != "(not used)":
        clean["impression_share"] = normalize_impression_share(df[is_col])
    else:
        clean["impression_share"] = np.nan

    cat_col = mapping.get("category")
    if cat_col and cat_col != "(not used)":
        clean["category_from_file"] = df[cat_col].astype(str).str.strip().replace({"": "Unassigned"})
    else:
        clean["category_from_file"] = "Unassigned"

    clean = clean.replace([np.inf, -np.inf], np.nan)
    clean = clean.dropna(subset=["date", "campaign_key", "campaign_name"])
    clean = clean[(clean["cost"] >= 0) & (clean["conversions"] >= 0)]
    clean["week_start"] = monday_of_week(clean["date"])
    clean["ctr"] = np.where(clean["impressions"] > 0, clean["clicks"] / clean["impressions"], np.nan)
    clean["cvr"] = np.where(clean["clicks"] > 0, clean["conversions"] / clean["clicks"], np.nan)
    clean["cpa"] = np.where(clean["conversions"] > 0, clean["cost"] / clean["conversions"], np.nan)
    return clean.sort_values(["date", "campaign_name", "campaign_key"]).reset_index(drop=True)



def initialize_campaign_mapping(clean_df):
    unique_campaigns = (
        clean_df[["campaign_key", "campaign_name", "category_from_file"]]
        .drop_duplicates()
        .sort_values(["campaign_name", "campaign_key"])
        .reset_index(drop=True)
    )
    mapping_df = unique_campaigns.rename(columns={"category_from_file": "category"}).copy()
    mapping_df["active"] = True
    mapping_df["category"] = mapping_df["category"].replace({"nan": "Unassigned"}).fillna("Unassigned")
    return mapping_df



def ensure_campaign_mapping(clean_df):
    campaign_signature = tuple(
        clean_df[["campaign_key", "campaign_name"]]
        .drop_duplicates()
        .sort_values(["campaign_name", "campaign_key"])
        .itertuples(index=False, name=None)
    )

    if st.session_state.get("campaign_signature") != campaign_signature:
        st.session_state["campaign_signature"] = campaign_signature
        st.session_state["campaign_mapping"] = initialize_campaign_mapping(clean_df)

    edited = st.data_editor(
        st.session_state["campaign_mapping"],
        hide_index=True,
        width="stretch",
        key="campaign_mapping_editor",
        disabled=["campaign_key", "campaign_name"],
    )
    st.session_state["campaign_mapping"] = edited
    return edited



def apply_campaign_mapping(clean_df, campaign_mapping):
    mapped = clean_df.merge(
        campaign_mapping[["campaign_key", "category", "active"]],
        on="campaign_key",
        how="left",
    )
    mapped["category"] = mapped["category"].fillna("Unassigned").replace({"": "Unassigned"})
    mapped["active"] = mapped["active"].fillna(True)
    mapped = mapped[mapped["active"] == True].copy()
    return mapped



def add_week_metadata(df, week_col="week_start"):
    out = df.copy()
    out[week_col] = pd.to_datetime(out[week_col])
    iso = out[week_col].dt.isocalendar()
    out["iso_year"] = iso.year.astype(int)
    out["iso_week"] = iso.week.astype(int)
    out["week_label"] = out.apply(lambda r: f"Week {int(r['iso_week'])} {int(r['iso_year'])}", axis=1)
    return out



def aggregate_weekly_by_campaign(mapped_df):
    agg_dict = {
        "cost": "sum",
        "conversions": "sum",
        "clicks": "sum",
        "impressions": "sum",
        "conversion_value": "sum",
        "impression_share": "mean",
    }
    weekly = (
        mapped_df.groupby(["week_start", "category", "campaign_key", "campaign_name"], as_index=False)
        .agg(agg_dict)
        .rename(columns={"cost": "spend"})
    )
    weekly["cpa"] = np.where(weekly["conversions"] > 0, weekly["spend"] / weekly["conversions"], np.nan)
    weekly["roas"] = np.where(weekly["spend"] > 0, weekly["conversion_value"] / weekly["spend"], np.nan)
    return add_week_metadata(weekly).sort_values(["week_start", "category", "campaign_name", "campaign_key"]).reset_index(drop=True)



def aggregate_weekly_by_category(campaign_weekly_df):
    weekly = (
        campaign_weekly_df.groupby(["week_start", "category"], as_index=False)
        .agg(
            spend=("spend", "sum"),
            conversions=("conversions", "sum"),
            clicks=("clicks", "sum"),
            impressions=("impressions", "sum"),
            conversion_value=("conversion_value", "sum"),
            impression_share=("impression_share", "mean"),
        )
    )
    weekly["cpa"] = np.where(weekly["conversions"] > 0, weekly["spend"] / weekly["conversions"], np.nan)
    weekly["roas"] = np.where(weekly["spend"] > 0, weekly["conversion_value"] / weekly["spend"], np.nan)
    return add_week_metadata(weekly).sort_values(["week_start", "category"]).reset_index(drop=True)



def default_max_budget(recent_avg_spend, recent_impression_share):
    if pd.isna(recent_avg_spend):
        recent_avg_spend = 0.0
    if pd.isna(recent_impression_share):
        multiplier = 2.0
    else:
        multiplier = np.clip(1.1 + (1.0 - recent_impression_share) * 2.0, 1.15, 3.0)
    suggested = max(50.0, recent_avg_spend * multiplier)
    return float(suggested)



def build_campaign_constraints(campaign_weekly_df):
    if campaign_weekly_df.empty:
        return pd.DataFrame(
            columns=[
                "campaign_key",
                "campaign_name",
                "category",
                "active",
                "recent_avg_weekly_spend",
                "recent_avg_weekly_conversions",
                "recent_avg_impression_share",
                "min_weekly_budget",
                "max_weekly_budget",
            ]
        )

    latest_week = campaign_weekly_df["week_start"].max()
    recent_cutoff = latest_week - pd.Timedelta(weeks=7)
    recent = campaign_weekly_df[campaign_weekly_df["week_start"] >= recent_cutoff].copy()

    summary = (
        recent.groupby(["campaign_key", "campaign_name", "category"], as_index=False)
        .agg(
            recent_avg_weekly_spend=("spend", "mean"),
            recent_avg_weekly_conversions=("conversions", "mean"),
            recent_avg_impression_share=("impression_share", "mean"),
        )
        .sort_values(["category", "campaign_name", "campaign_key"])
        .reset_index(drop=True)
    )
    summary["active"] = True
    summary["min_weekly_budget"] = 0.0
    summary["max_weekly_budget"] = summary.apply(
        lambda row: default_max_budget(row["recent_avg_weekly_spend"], row["recent_avg_impression_share"]),
        axis=1,
    )
    return summary



def ensure_campaign_constraints(campaign_weekly_df):
    signature = tuple(
        campaign_weekly_df[["campaign_key", "campaign_name", "category"]]
        .drop_duplicates()
        .sort_values(["category", "campaign_name", "campaign_key"])
        .itertuples(index=False, name=None)
    )
    if st.session_state.get("constraint_signature") != signature:
        st.session_state["constraint_signature"] = signature
        st.session_state["campaign_constraints"] = build_campaign_constraints(campaign_weekly_df)

    constraints = st.data_editor(
        st.session_state["campaign_constraints"],
        hide_index=True,
        width="stretch",
        key="campaign_constraints_editor",
        disabled=[
            "campaign_key",
            "campaign_name",
            "category",
            "recent_avg_weekly_spend",
            "recent_avg_weekly_conversions",
            "recent_avg_impression_share",
        ],
    )
    constraints["min_weekly_budget"] = pd.to_numeric(constraints["min_weekly_budget"], errors="coerce").fillna(0.0)
    constraints["max_weekly_budget"] = pd.to_numeric(constraints["max_weekly_budget"], errors="coerce").fillna(0.0)
    constraints["max_weekly_budget"] = np.maximum(constraints["max_weekly_budget"], constraints["min_weekly_budget"])
    st.session_state["campaign_constraints"] = constraints
    return constraints



def build_category_benchmarks(campaign_weekly_df):
    positive = campaign_weekly_df[campaign_weekly_df["spend"] > 0].copy()
    if positive.empty:
        return {"__overall__": {"median_efficiency": 0.01, "median_spend": 100.0, "max_conversions": 1.0}}

    positive["efficiency"] = positive["conversions"] / positive["spend"].replace(0, np.nan)
    overall = {
        "median_efficiency": float(max(positive["efficiency"].median(skipna=True), 0.01)),
        "median_spend": float(max(positive["spend"].median(skipna=True), 100.0)),
        "max_conversions": float(max(positive["conversions"].max(skipna=True), 1.0)),
    }

    benchmarks = {"__overall__": overall}
    by_category = positive.groupby("category")
    for category, grp in by_category:
        eff = grp["efficiency"].median(skipna=True)
        median_spend = grp["spend"].median(skipna=True)
        max_conversions = grp["conversions"].max(skipna=True)
        benchmarks[category] = {
            "median_efficiency": float(max(eff if not pd.isna(eff) else overall["median_efficiency"], 0.01)),
            "median_spend": float(max(median_spend if not pd.isna(median_spend) else overall["median_spend"], 25.0)),
            "max_conversions": float(max(max_conversions if not pd.isna(max_conversions) else overall["max_conversions"], 1.0)),
        }
    return benchmarks



def heuristic_model(campaign_key, campaign_name, category, campaign_df, benchmarks, fit_status):
    benchmark = benchmarks.get(category, benchmarks["__overall__"])
    positive = campaign_df[campaign_df["spend"] > 0].copy()

    recent_avg_spend = float(campaign_df["spend"].tail(8).mean()) if not campaign_df.empty else 0.0
    recent_avg_conversions = float(campaign_df["conversions"].tail(8).mean()) if not campaign_df.empty else 0.0
    points = len(campaign_df)

    if positive.empty:
        median_efficiency = benchmark["median_efficiency"]
        median_spend = benchmark["median_spend"]
        max_conversions = benchmark["max_conversions"]
    else:
        efficiencies = (positive["conversions"] / positive["spend"]).replace([np.inf, -np.inf], np.nan)
        median_efficiency = efficiencies.median(skipna=True)
        if pd.isna(median_efficiency) or median_efficiency <= 0:
            median_efficiency = benchmark["median_efficiency"]
        median_spend = positive["spend"].median(skipna=True)
        if pd.isna(median_spend) or median_spend <= 0:
            median_spend = benchmark["median_spend"]
        max_conversions = positive["conversions"].max(skipna=True)
        if pd.isna(max_conversions) or max_conversions <= 0:
            max_conversions = benchmark["max_conversions"]

    a = max(float(max_conversions) * 1.5, float(median_efficiency) * float(median_spend) * 2.0, 1.0)
    b = max(float(median_spend), 25.0)

    return CampaignModel(
        campaign_key=campaign_key,
        campaign_name=campaign_name,
        category=category,
        a=float(a),
        b=float(b),
        fit_status=fit_status,
        data_points=int(points),
        recent_avg_spend=float(recent_avg_spend),
        recent_avg_conversions=float(recent_avg_conversions),
    )



def fit_single_campaign_model(campaign_df, benchmarks):
    campaign_df = campaign_df.copy().sort_values("week_start")
    campaign_df = campaign_df[(campaign_df["spend"] >= 0) & (campaign_df["conversions"] >= 0)]
    campaign_df = campaign_df[["campaign_key", "campaign_name", "category", "spend", "conversions", "week_start"]].dropna()

    campaign_key = str(campaign_df["campaign_key"].iloc[0])
    campaign_name = str(campaign_df["campaign_name"].iloc[0])
    category = str(campaign_df["category"].iloc[0])

    if len(campaign_df) < 6 or campaign_df["spend"].nunique() < 4:
        return heuristic_model(campaign_key, campaign_name, category, campaign_df, benchmarks, fit_status="heuristic: limited history")

    x = campaign_df["spend"].astype(float).values
    y = campaign_df["conversions"].astype(float).values

    x_positive = x[x > 0]
    benchmark = benchmarks.get(category, benchmarks["__overall__"])
    median_x = float(np.median(x_positive)) if len(x_positive) else float(benchmark["median_spend"])
    max_y = max(float(np.nanmax(y)), 1.0)
    mean_efficiency = float(np.nanmean(np.where(x > 0, y / np.maximum(x, 1e-9), np.nan)))
    if np.isnan(mean_efficiency) or mean_efficiency <= 0:
        mean_efficiency = float(benchmark["median_efficiency"])

    initial_b = max(median_x, 25.0)
    initial_a = max(max_y * 1.5, mean_efficiency * initial_b * 2.0, 1.0)

    upper_a = max(max_y * 10.0, initial_a * 5.0, 10.0)
    upper_b = max(float(np.nanmax(x)) * 10.0, initial_b * 10.0, 250.0)

    try:
        params, _ = curve_fit(
            hill_curve,
            x,
            y,
            p0=[initial_a, initial_b],
            bounds=([0.01, 0.01], [upper_a, upper_b]),
            maxfev=30000,
        )
        a, b = float(params[0]), float(params[1])
        preds = hill_curve(x, a, b)
        mae = np.nanmean(np.abs(preds - y))
        if np.any(np.isnan(preds)) or mae > max(5.0, np.nanmean(y) * 3.0):
            return heuristic_model(campaign_key, campaign_name, category, campaign_df, benchmarks, fit_status="heuristic: unstable fit")
        return CampaignModel(
            campaign_key=campaign_key,
            campaign_name=campaign_name,
            category=category,
            a=a,
            b=b,
            fit_status="curve_fit",
            data_points=len(campaign_df),
            recent_avg_spend=float(campaign_df["spend"].tail(8).mean()),
            recent_avg_conversions=float(campaign_df["conversions"].tail(8).mean()),
        )
    except Exception:
        return heuristic_model(campaign_key, campaign_name, category, campaign_df, benchmarks, fit_status="heuristic: fit failed")



def fit_all_campaign_models(campaign_weekly_df, campaign_constraints):
    active_campaigns = campaign_constraints[campaign_constraints["active"] == True]["campaign_key"].tolist()
    models: Dict[str, CampaignModel] = {}
    diagnostics = []
    benchmarks = build_category_benchmarks(campaign_weekly_df)

    for campaign_key in active_campaigns:
        campaign_df = campaign_weekly_df[campaign_weekly_df["campaign_key"] == campaign_key].copy()
        if campaign_df.empty:
            continue
        model = fit_single_campaign_model(campaign_df, benchmarks)
        models[campaign_key] = model
        diagnostics.append(
            {
                "campaign_key": model.campaign_key,
                "campaign_name": model.campaign_name,
                "category": model.category,
                "fit_status": model.fit_status,
                "data_points": model.data_points,
                "curve_a": round(model.a, 4),
                "curve_b": round(model.b, 4),
                "recent_avg_weekly_spend": round(model.recent_avg_spend, 2),
                "recent_avg_weekly_conversions": round(model.recent_avg_conversions, 2),
            }
        )
    return models, pd.DataFrame(diagnostics)



def compute_month_factor_map(weekly_df, entity_col, min_total_weeks=10, min_month_obs=2):
    if weekly_df.empty:
        return {}
    source = (
        weekly_df.groupby([entity_col, "week_start"], as_index=False)
        .agg(spend=("spend", "sum"), conversions=("conversions", "sum"))
        .sort_values([entity_col, "week_start"])
    )
    source = source[source["spend"] > 0].copy()
    if source.empty:
        return {}
    source["efficiency"] = source["conversions"] / source["spend"].replace(0, np.nan)
    source["month"] = pd.to_datetime(source["week_start"]).dt.month

    factor_map = {}
    for entity, grp in source.groupby(entity_col):
        if len(grp) < min_total_weeks:
            continue
        baseline = grp["efficiency"].median(skipna=True)
        if pd.isna(baseline) or baseline <= 0:
            continue
        for month in range(1, 13):
            sample = grp.loc[grp["month"] == month, "efficiency"]
            if len(sample) < min_month_obs:
                continue
            factor = float(np.clip(sample.median(skipna=True) / baseline, 0.7, 1.3))
            factor_map[(str(entity), int(month))] = factor
    return factor_map



def seasonality_factor_for_campaign(campaign_key, category, forecast_week, campaign_factor_map, category_factor_map):
    month = int(pd.Timestamp(forecast_week).month)
    return float(
        campaign_factor_map.get((str(campaign_key), month), category_factor_map.get((str(category), month), 1.0))
    )



def campaign_meta_for_optimizer(campaign_constraints, models):
    active = campaign_constraints[campaign_constraints["active"] == True].copy()
    active = active[active["campaign_key"].isin(models.keys())].copy()
    active = active.sort_values(["category", "campaign_name", "campaign_key"]).reset_index(drop=True)
    return active



def validate_budget_feasibility(total_budget, mins, maxs):
    min_needed = float(np.sum(mins))
    max_allowed = float(np.sum(maxs))
    if total_budget < min_needed - 1e-6:
        raise ValueError(f"Budget is below the required campaign minimums. Minimum needed: {min_needed:,.2f}")
    if total_budget > max_allowed + 1e-6:
        raise ValueError(f"Budget is above the allowed campaign maximums. Maximum allowed: {max_allowed:,.2f}")



def greedy_campaign_allocate(total_budget, campaign_meta, models, factor_map, step=None):
    campaign_keys = campaign_meta["campaign_key"].tolist()
    mins = campaign_meta["min_weekly_budget"].astype(float).to_numpy()
    maxs = campaign_meta["max_weekly_budget"].astype(float).to_numpy()
    validate_budget_feasibility(total_budget, mins, maxs)

    spend = mins.copy()
    remaining = float(total_budget - mins.sum())
    if remaining <= 1e-6:
        return spend

    if step is None:
        step = max(float(total_budget) / 400.0, 10.0)

    safety = 0
    while remaining > 1e-6 and safety < 200000:
        safety += 1
        best_idx = None
        best_gain = -1.0
        for idx, campaign_key in enumerate(campaign_keys):
            if spend[idx] >= maxs[idx] - 1e-6:
                continue
            delta = min(step, remaining, maxs[idx] - spend[idx])
            factor = factor_map.get(campaign_key, 1.0)
            current_pred = models[campaign_key].predict(spend[idx], factor)
            new_pred = models[campaign_key].predict(spend[idx] + delta, factor)
            gain = (new_pred - current_pred) / max(delta, 1e-9)
            if gain > best_gain:
                best_gain = gain
                best_idx = idx
        if best_idx is None:
            break
        delta = min(step, remaining, maxs[best_idx] - spend[best_idx])
        spend[best_idx] += delta
        remaining -= delta

    if remaining > 1e-3:
        # Small cleanup pass into any remaining capacity.
        for idx in range(len(campaign_keys)):
            if remaining <= 1e-6:
                break
            room = maxs[idx] - spend[idx]
            if room <= 1e-6:
                continue
            delta = min(room, remaining)
            spend[idx] += delta
            remaining -= delta

    if remaining > 1e-3:
        raise ValueError("Could not allocate the full budget within the campaign constraints.")
    return spend



def solve_optimal_campaign_mix(total_budget, campaign_meta, models, factor_map):
    campaign_keys = campaign_meta["campaign_key"].tolist()
    mins = campaign_meta["min_weekly_budget"].astype(float).to_numpy()
    maxs = campaign_meta["max_weekly_budget"].astype(float).to_numpy()
    validate_budget_feasibility(total_budget, mins, maxs)

    x0 = greedy_campaign_allocate(total_budget, campaign_meta, models, factor_map)

    def objective(x):
        return -float(sum(models[key].predict(x[idx], factor_map.get(key, 1.0)) for idx, key in enumerate(campaign_keys)))

    def gradient(x):
        return -np.array([models[key].marginal(x[idx], factor_map.get(key, 1.0)) for idx, key in enumerate(campaign_keys)], dtype=float)

    constraints = [{"type": "eq", "fun": lambda x: np.sum(x) - float(total_budget), "jac": lambda x: np.ones_like(x)}]
    bounds = list(zip(mins, maxs))

    try:
        result = minimize(
            objective,
            x0=x0,
            jac=gradient,
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
            options={"maxiter": 400, "ftol": 1e-8, "disp": False},
        )
        if result.success and np.isfinite(result.x).all():
            x = np.clip(result.x, mins, maxs)
            spend_gap = total_budget - float(np.sum(x))
            if abs(spend_gap) > 1e-5:
                x = greedy_campaign_allocate(total_budget, campaign_meta, models, factor_map, step=max(total_budget / 1000.0, 1.0))
            return x
    except Exception:
        pass

    return x0



def predicted_total_conversions(total_budget, campaign_meta, models, factor_map):
    spend = solve_optimal_campaign_mix(total_budget, campaign_meta, models, factor_map)
    campaign_keys = campaign_meta["campaign_key"].tolist()
    total = sum(models[key].predict(spend[idx], factor_map.get(key, 1.0)) for idx, key in enumerate(campaign_keys))
    return float(total), spend



def solve_budget_for_goal(goal_conversions, campaign_meta, models, factor_map):
    mins = campaign_meta["min_weekly_budget"].astype(float).to_numpy()
    maxs = campaign_meta["max_weekly_budget"].astype(float).to_numpy()

    low = float(mins.sum())
    high = float(maxs.sum())

    max_possible, _ = predicted_total_conversions(high, campaign_meta, models, factor_map)
    if goal_conversions > max_possible:
        raise ValueError(
            "Your weekly goal is higher than this version of the model thinks is possible inside the current max budgets. "
            f"Estimated max weekly conversions: {max_possible:,.2f}"
        )

    best_budget = high
    best_spend = None
    for _ in range(35):
        mid = (low + high) / 2.0
        convs, spend = predicted_total_conversions(mid, campaign_meta, models, factor_map)
        if convs >= goal_conversions:
            best_budget = mid
            best_spend = spend
            high = mid
        else:
            low = mid

    if best_spend is None:
        _, best_spend = predicted_total_conversions(best_budget, campaign_meta, models, factor_map)
    return float(best_budget), best_spend



def compute_factor_map_for_week(campaign_meta, forecast_week, campaign_factor_map, category_factor_map):
    return {
        row["campaign_key"]: seasonality_factor_for_campaign(
            row["campaign_key"],
            row["category"],
            forecast_week,
            campaign_factor_map,
            category_factor_map,
        )
        for _, row in campaign_meta.iterrows()
    }



def build_prior_year_lookup(campaign_weekly_df):
    campaign_hist = campaign_weekly_df[[
        "campaign_key", "campaign_name", "category", "week_start", "iso_year", "iso_week", "week_label", "spend", "conversions"
    ]].copy()
    category_hist = (
        campaign_weekly_df.groupby(["category", "week_start", "iso_year", "iso_week", "week_label"], as_index=False)
        .agg(spend=("spend", "sum"), conversions=("conversions", "sum"))
    )
    portfolio_hist = (
        campaign_weekly_df.groupby(["week_start", "iso_year", "iso_week", "week_label"], as_index=False)
        .agg(spend=("spend", "sum"), conversions=("conversions", "sum"))
    )
    return campaign_hist, category_hist, portfolio_hist



def attach_prior_year_comparison(df, prior_df, join_keys, spend_col_name, conv_col_name):
    if df.empty:
        return df.copy()
    out = df.copy()
    out["compare_iso_year"] = out["iso_year"] - 1
    prior = prior_df.copy()
    prior = prior.rename(columns={"spend": spend_col_name, "conversions": conv_col_name, "week_label": "prior_year_week_label"})
    merge_keys_left = join_keys + ["compare_iso_year", "iso_week"]
    merge_keys_right = join_keys + ["iso_year", "iso_week"]
    out = out.merge(
        prior[merge_keys_right + [spend_col_name, conv_col_name, "prior_year_week_label"]],
        left_on=merge_keys_left,
        right_on=merge_keys_right,
        how="left",
    )
    out = out.drop(columns=["compare_iso_year", "iso_year_y"], errors="ignore")
    if "iso_year_x" in out.columns:
        out = out.rename(columns={"iso_year_x": "iso_year"})
    return out



def build_forecast(campaign_weekly_df, campaign_constraints, models, horizon_weeks, input_mode, weekly_budget=None, weekly_goal=None):
    campaign_meta = campaign_meta_for_optimizer(campaign_constraints, models)
    campaign_factor_map = compute_month_factor_map(campaign_weekly_df, "campaign_key", min_total_weeks=10, min_month_obs=2)
    category_factor_map = compute_month_factor_map(campaign_weekly_df, "category", min_total_weeks=8, min_month_obs=2)
    prior_campaign, prior_category, prior_portfolio = build_prior_year_lookup(campaign_weekly_df)

    if campaign_meta.empty:
        raise ValueError("No active campaigns are available for optimization.")

    start_week = pd.Timestamp(campaign_weekly_df["week_start"].max()) + pd.Timedelta(weeks=1)

    detail_rows = []
    summary_rows = []
    category_rows = []

    for week_index in range(horizon_weeks):
        forecast_week = start_week + pd.Timedelta(weeks=week_index)
        factor_map = compute_factor_map_for_week(campaign_meta, forecast_week, campaign_factor_map, category_factor_map)

        if input_mode == "I know my weekly budget":
            if weekly_budget is None:
                raise ValueError("Weekly budget is required.")
            spend_vector = solve_optimal_campaign_mix(weekly_budget, campaign_meta, models, factor_map)
            chosen_budget = float(weekly_budget)
        else:
            if weekly_goal is None:
                raise ValueError("Weekly goal is required.")
            chosen_budget, spend_vector = solve_budget_for_goal(weekly_goal, campaign_meta, models, factor_map)

        week_total_conversions = 0.0
        week_category_totals: Dict[str, Dict[str, float]] = {}

        for idx, row in campaign_meta.iterrows():
            campaign_key = row["campaign_key"]
            campaign_name = row["campaign_name"]
            category = row["category"]
            recommended_budget = float(spend_vector[idx])
            factor = float(factor_map.get(campaign_key, 1.0))
            predicted_conversions = float(models[campaign_key].predict(recommended_budget, factor))
            marginal_conv_per_100 = float(models[campaign_key].marginal(recommended_budget, factor) * 100.0)
            week_total_conversions += predicted_conversions

            detail_rows.append(
                {
                    "week_start": forecast_week.normalize(),
                    "campaign_key": campaign_key,
                    "campaign_name": campaign_name,
                    "category": category,
                    "recommended_budget": round(recommended_budget, 2),
                    "predicted_conversions": round(predicted_conversions, 2),
                    "marginal_conversions_per_100": round(marginal_conv_per_100, 3),
                    "recent_avg_weekly_spend": round(float(row["recent_avg_weekly_spend"]), 2),
                    "budget_change_vs_recent_avg": round(recommended_budget - float(row["recent_avg_weekly_spend"]), 2),
                    "seasonality_factor": round(factor, 3),
                }
            )

            bucket = week_category_totals.setdefault(category, {"budget": 0.0, "conversions": 0.0})
            bucket["budget"] += recommended_budget
            bucket["conversions"] += predicted_conversions

        summary_rows.append(
            {
                "week_start": forecast_week.normalize(),
                "recommended_total_budget": round(chosen_budget, 2),
                "predicted_total_conversions": round(week_total_conversions, 2),
            }
        )

        for category, values in week_category_totals.items():
            category_rows.append(
                {
                    "week_start": forecast_week.normalize(),
                    "category": category,
                    "recommended_budget": round(values["budget"], 2),
                    "predicted_conversions": round(values["conversions"], 2),
                }
            )

    detail_df = add_week_metadata(pd.DataFrame(detail_rows))
    category_df = add_week_metadata(pd.DataFrame(category_rows))
    summary_df = add_week_metadata(pd.DataFrame(summary_rows))

    detail_df = attach_prior_year_comparison(
        detail_df,
        prior_campaign,
        join_keys=["campaign_key"],
        spend_col_name="prior_year_actual_spend",
        conv_col_name="prior_year_actual_conversions",
    )
    category_df = attach_prior_year_comparison(
        category_df,
        prior_category,
        join_keys=["category"],
        spend_col_name="prior_year_actual_spend",
        conv_col_name="prior_year_actual_conversions",
    )
    summary_df = attach_prior_year_comparison(
        summary_df,
        prior_portfolio,
        join_keys=[],
        spend_col_name="prior_year_actual_total_spend",
        conv_col_name="prior_year_actual_total_conversions",
    )

    return detail_df.sort_values(["week_start", "category", "campaign_name"]), category_df.sort_values(["week_start", "category"]), summary_df.sort_values("week_start")



def build_next_best_dollar_table(campaign_constraints, models, campaign_weekly_df, reference_week):
    campaign_meta = campaign_meta_for_optimizer(campaign_constraints, models)
    campaign_factor_map = compute_month_factor_map(campaign_weekly_df, "campaign_key", min_total_weeks=10, min_month_obs=2)
    category_factor_map = compute_month_factor_map(campaign_weekly_df, "category", min_total_weeks=8, min_month_obs=2)
    factor_map = compute_factor_map_for_week(campaign_meta, reference_week, campaign_factor_map, category_factor_map)

    rows = []
    for _, row in campaign_meta.iterrows():
        campaign_key = row["campaign_key"]
        baseline_spend = float(row["recent_avg_weekly_spend"])
        factor = float(factor_map.get(campaign_key, 1.0))
        rows.append(
            {
                "campaign_key": campaign_key,
                "campaign_name": row["campaign_name"],
                "category": row["category"],
                "recent_avg_weekly_spend": round(baseline_spend, 2),
                "marginal_conversions_per_100_at_recent_spend": round(models[campaign_key].marginal(baseline_spend, factor) * 100.0, 3),
                "seasonality_factor": round(factor, 3),
                "max_weekly_budget": round(float(row["max_weekly_budget"]), 2),
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values("marginal_conversions_per_100_at_recent_spend", ascending=False).reset_index(drop=True)



def build_historical_audit(campaign_weekly_df, campaign_constraints, models):
    if campaign_weekly_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    active_constraints = campaign_constraints[campaign_constraints["active"] == True].copy()
    if active_constraints.empty:
        return pd.DataFrame(), pd.DataFrame()

    campaign_factor_map = compute_month_factor_map(campaign_weekly_df, "campaign_key", min_total_weeks=10, min_month_obs=2)
    category_factor_map = compute_month_factor_map(campaign_weekly_df, "category", min_total_weeks=8, min_month_obs=2)
    first_seen = campaign_weekly_df.groupby("campaign_key", as_index=False)["week_start"].min().rename(columns={"week_start": "first_seen_week"})

    summary_rows = []
    detail_rows = []

    for week in sorted(campaign_weekly_df["week_start"].dropna().unique()):
        week_ts = pd.Timestamp(week)
        actual_week_df = campaign_weekly_df[campaign_weekly_df["week_start"] == week_ts].copy()
        if actual_week_df.empty:
            continue

        eligible_keys = first_seen.loc[first_seen["first_seen_week"] <= week_ts, "campaign_key"].astype(str).tolist()
        week_constraints = active_constraints[active_constraints["campaign_key"].astype(str).isin(eligible_keys)].copy()
        week_constraints = week_constraints[week_constraints["campaign_key"].isin(models.keys())].copy()
        if week_constraints.empty:
            continue

        campaign_meta = campaign_meta_for_optimizer(week_constraints, models)
        if campaign_meta.empty:
            continue

        factor_map = compute_factor_map_for_week(campaign_meta, week_ts, campaign_factor_map, category_factor_map)
        actual_total_budget = float(actual_week_df["spend"].sum())

        mins = campaign_meta["min_weekly_budget"].astype(float).to_numpy()
        maxs = campaign_meta["max_weekly_budget"].astype(float).to_numpy()
        if actual_total_budget < float(mins.sum()) - 1e-6:
            actual_total_budget = float(mins.sum())
        if actual_total_budget > float(maxs.sum()) + 1e-6:
            actual_total_budget = float(maxs.sum())

        optimal_spend = solve_optimal_campaign_mix(actual_total_budget, campaign_meta, models, factor_map)

        actual_spend_lookup = actual_week_df.groupby("campaign_key", as_index=False).agg(
            actual_spend=("spend", "sum"),
            actual_conversions=("conversions", "sum"),
        )
        actual_spend_lookup["campaign_key"] = actual_spend_lookup["campaign_key"].astype(str)
        actual_lookup = actual_spend_lookup.set_index("campaign_key").to_dict(orient="index")

        model_est_actual = 0.0
        model_est_opt = 0.0
        actual_conversions_total = float(actual_week_df["conversions"].sum())

        for idx, row in campaign_meta.iterrows():
            campaign_key = str(row["campaign_key"])
            factor = float(factor_map.get(campaign_key, 1.0))
            actual_spend = float(actual_lookup.get(campaign_key, {}).get("actual_spend", 0.0))
            actual_conversions = float(actual_lookup.get(campaign_key, {}).get("actual_conversions", 0.0))
            predicted_at_actual = float(models[campaign_key].predict(actual_spend, factor))
            predicted_at_optimal = float(models[campaign_key].predict(optimal_spend[idx], factor))
            model_est_actual += predicted_at_actual
            model_est_opt += predicted_at_optimal

            detail_rows.append(
                {
                    "week_start": week_ts.normalize(),
                    "campaign_key": campaign_key,
                    "campaign_name": row["campaign_name"],
                    "category": row["category"],
                    "actual_spend": round(actual_spend, 2),
                    "optimal_spend": round(float(optimal_spend[idx]), 2),
                    "budget_delta": round(float(optimal_spend[idx]) - actual_spend, 2),
                    "actual_conversions": round(actual_conversions, 2),
                    "model_estimated_conversions_at_actual_spend": round(predicted_at_actual, 2),
                    "model_estimated_conversions_at_optimal_spend": round(predicted_at_optimal, 2),
                    "seasonality_factor": round(factor, 3),
                }
            )

        summary_rows.append(
            {
                "week_start": week_ts.normalize(),
                "actual_total_budget": round(actual_total_budget, 2),
                "actual_total_conversions": round(actual_conversions_total, 2),
                "model_estimated_conversions_at_actual_mix": round(model_est_actual, 2),
                "model_estimated_conversions_at_optimal_mix": round(model_est_opt, 2),
                "modeled_conversion_lift": round(model_est_opt - model_est_actual, 2),
                "modeled_conversion_lift_pct": round(((model_est_opt / model_est_actual) - 1.0) if model_est_actual > 0 else np.nan, 4),
            }
        )

    summary_df = add_week_metadata(pd.DataFrame(summary_rows)) if summary_rows else pd.DataFrame()
    detail_df = add_week_metadata(pd.DataFrame(detail_rows)) if detail_rows else pd.DataFrame()
    return summary_df.sort_values("week_start"), detail_df.sort_values(["week_start", "category", "campaign_name"])



def csv_download_bytes(df):
    return df.to_csv(index=False).encode("utf-8")



def apply_excel_formatting(writer, sheet_name, df):
    ws = writer.book[sheet_name]
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for idx, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].head(200).fillna(""))) if not df.empty else len(str(col))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)

        lowered = str(col).lower()
        if any(token in lowered for token in ["budget", "spend", "cost", "value"]):
            number_format = '$#,##0.00;[Red]($#,##0.00)'
        elif "pct" in lowered:
            number_format = '0.0%'
        elif any(token in lowered for token in ["factor"]):
            number_format = '0.000'
        elif any(token in lowered for token in ["conversions", "ctr", "cvr", "roas", "cpa"]):
            number_format = '0.00'
        elif "week_start" in lowered:
            number_format = 'yyyy-mm-dd'
        else:
            number_format = None

        if number_format:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                row[0].number_format = number_format



def build_excel_workbook_bytes(sheet_map):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for raw_name, df in sheet_map.items():
            safe_name = str(raw_name)[:31]
            export_df = df.copy()
            export_df.to_excel(writer, sheet_name=safe_name, index=False)
            apply_excel_formatting(writer, safe_name, export_df)
    buffer.seek(0)
    return buffer.getvalue()


st.title("Paid Search Mix Optimizer")
st.caption(
    "Upload a CSV, map your columns, assign campaign categories, and get campaign-level budget allocations that aim to invest the next best dollar for maximum conversions."
)

with st.expander("What this version does", expanded=False):
    st.markdown(
        """
        - Fits a diminishing-returns response curve to each active campaign.
        - Optimizes weekly budget at the **campaign** level, not just category level.
        - Rolls campaign recommendations back up into category mix views.
        - Exports Excel sheets with prior-year comparisons and a historical audit of what the optimizer would have recommended for the weeks in your file.
        - This is decision support, not autopilot. It assumes spend is the main driver and does not model auction shocks, creative changes, or tracking breaks.
        """
    )

uploaded_file = st.file_uploader("Upload your paid search CSV", type=["csv"])

if uploaded_file is None:
    st.info("Upload a CSV to begin. Use the sample file included with this project if you want to test the app first.")
    st.stop()

raw_df = read_uploaded_csv(uploaded_file)

st.subheader("1) Map your columns")
st.write("Choose the columns from your file that match each required field.")

columns = raw_df.columns.tolist()
optional_columns = ["(not used)"] + columns

mapping = {}
for field_key, label in REQUIRED_FIELDS:
    guessed = guess_column(field_key, columns, required=True)
    default_index = columns.index(guessed) if guessed in columns else 0
    mapping[field_key] = st.selectbox(label, options=columns, index=default_index, key=f"map_{field_key}")

for field_key, label in OPTIONAL_FIELDS:
    guessed = guess_column(field_key, columns, required=False)
    default_index = optional_columns.index(guessed) if guessed in optional_columns else 0
    mapping[field_key] = st.selectbox(label, options=optional_columns, index=default_index, key=f"map_{field_key}")

missing_required = [label for field_key, label in REQUIRED_FIELDS if mapping.get(field_key) in [None, "(not used)"]]
if missing_required:
    st.error("Map all required fields before continuing.")
    st.stop()

clean_df = standardize_dataset(raw_df, mapping)

if clean_df.empty:
    st.error("After cleaning, there are no usable rows left. Check your date and numeric columns.")
    st.stop()

col1, col2, col3 = st.columns(3)
col1.metric("Usable rows", f"{len(clean_df):,}")
col2.metric("Campaigns", f"{clean_df['campaign_key'].nunique():,}")
col3.metric("Date range", f"{clean_df['date'].min().date()} → {clean_df['date'].max().date()}")

with st.expander("Preview cleaned data", expanded=False):
    st.dataframe(clean_df.head(50), use_container_width=True)

st.subheader("2) Assign categories to campaigns")
st.write("Edit the category column below. Keep the names consistent, like Brand, Nonbrand, Competitor, Shopping.")

campaign_mapping = ensure_campaign_mapping(clean_df)
if campaign_mapping["category"].eq("Unassigned").any():
    st.warning("Some campaigns are still marked as Unassigned. That is okay for testing, but not ideal for a real allocation model.")

mapped_df = apply_campaign_mapping(clean_df, campaign_mapping)
campaign_weekly_df = aggregate_weekly_by_campaign(mapped_df)
category_weekly_df = aggregate_weekly_by_category(campaign_weekly_df)

if campaign_weekly_df.empty:
    st.error("No active data left after campaign mapping.")
    st.stop()

st.subheader("3) Set campaign budget rules")
st.write("These are the per-campaign floors and ceilings the optimizer must respect. The max budget suggestion loosely uses recent spend and impression share headroom.")

campaign_constraints = ensure_campaign_constraints(campaign_weekly_df)
active_campaigns = campaign_constraints[campaign_constraints["active"] == True]["campaign_key"].tolist()
if not active_campaigns:
    st.error("Turn on at least one campaign in the campaign budget rules table.")
    st.stop()

models, diagnostics_df = fit_all_campaign_models(campaign_weekly_df, campaign_constraints)
if not models:
    st.error("The app could not fit any active campaign models from the uploaded history.")
    st.stop()

st.subheader("4) Choose your planning mode")
input_mode = st.radio(
    "How do you want to plan?",
    options=["I know my weekly budget", "I have a weekly conversion goal"],
    horizontal=True,
)

horizon_weeks = int(st.number_input("Forecast horizon (weeks)", min_value=1, max_value=26, value=8, step=1))

weekly_budget = None
weekly_goal = None

if input_mode == "I know my weekly budget":
    recent_total_spend = float(
        campaign_weekly_df.groupby("week_start")["spend"].sum().tail(8).mean()
    ) if not campaign_weekly_df.empty else 1000.0
    weekly_budget = float(
        st.number_input(
            "Weekly total budget",
            min_value=0.0,
            value=round(max(recent_total_spend, 100.0), 2),
            step=100.0,
        )
    )
else:
    recent_total_conversions = float(
        campaign_weekly_df.groupby("week_start")["conversions"].sum().tail(8).mean()
    ) if not campaign_weekly_df.empty else 50.0
    weekly_goal = float(
        st.number_input(
            "Weekly conversion goal",
            min_value=0.0,
            value=round(max(recent_total_conversions, 1.0), 2),
            step=5.0,
        )
    )

try:
    detail_forecast, category_forecast, summary_forecast = build_forecast(
        campaign_weekly_df=campaign_weekly_df,
        campaign_constraints=campaign_constraints,
        models=models,
        horizon_weeks=horizon_weeks,
        input_mode=input_mode,
        weekly_budget=weekly_budget,
        weekly_goal=weekly_goal,
    )
    audit_summary, audit_detail = build_historical_audit(
        campaign_weekly_df=campaign_weekly_df,
        campaign_constraints=campaign_constraints,
        models=models,
    )
except ValueError as exc:
    st.error(str(exc))
    st.stop()

reference_week = pd.Timestamp(summary_forecast["week_start"].min()) if not summary_forecast.empty else pd.Timestamp(campaign_weekly_df["week_start"].max()) + pd.Timedelta(weeks=1)
next_best_dollars = build_next_best_dollar_table(campaign_constraints, models, campaign_weekly_df, reference_week)

st.subheader("Forecast summary")
st.dataframe(summary_forecast, use_container_width=True)

summary_chart = summary_forecast.set_index("week_label")[["recommended_total_budget", "predicted_total_conversions"]]
st.line_chart(summary_chart)

st.subheader("Recommended category mix by week")
st.dataframe(category_forecast, use_container_width=True)
category_pivot = category_forecast.pivot(index="week_label", columns="category", values="recommended_budget")
st.bar_chart(category_pivot)

st.subheader("Recommended campaign mix by week")
st.dataframe(detail_forecast, use_container_width=True)

st.subheader("Next best dollars")
st.write("This ranks campaigns by the estimated incremental conversions from the next $100 starting from recent average spend.")
st.dataframe(next_best_dollars, use_container_width=True)

st.subheader("Historical audit")
st.write("This is a retrospective efficiency check. It compares each historical week's actual mix to what the optimizer would have recommended using that week's total spend. It is not a true walk-forward backtest.")
st.dataframe(audit_summary, use_container_width=True)

with st.expander("Historical audit by campaign", expanded=False):
    st.dataframe(audit_detail, use_container_width=True)

with st.expander("Model diagnostics", expanded=False):
    st.dataframe(diagnostics_df, use_container_width=True)
    st.write("Weekly campaign history used for modeling")
    st.dataframe(campaign_weekly_df, use_container_width=True)
    st.write("Weekly category rollup")
    st.dataframe(category_weekly_df, use_container_width=True)

st.download_button(
    "Download campaign forecast CSV",
    data=csv_download_bytes(detail_forecast),
    file_name="campaign_mix_forecast.csv",
    mime="text/csv",
)

st.download_button(
    "Download category forecast CSV",
    data=csv_download_bytes(category_forecast),
    file_name="category_mix_forecast.csv",
    mime="text/csv",
)

st.download_button(
    "Download forecast summary CSV",
    data=csv_download_bytes(summary_forecast),
    file_name="forecast_summary.csv",
    mime="text/csv",
)

st.download_button(
    "Download historical audit summary CSV",
    data=csv_download_bytes(audit_summary),
    file_name="historical_audit_summary.csv",
    mime="text/csv",
)

st.download_button(
    "Download historical audit detail CSV",
    data=csv_download_bytes(audit_detail),
    file_name="historical_audit_detail.csv",
    mime="text/csv",
)

excel_payload = build_excel_workbook_bytes(
    {
        "forecast_summary": summary_forecast,
        "category_forecast": category_forecast,
        "campaign_forecast": detail_forecast,
        "next_best_dollars": next_best_dollars,
        "historical_audit_summary": audit_summary,
        "historical_audit_detail": audit_detail,
        "campaign_model_diagnostics": diagnostics_df,
        "campaign_history": campaign_weekly_df,
        "category_history": category_weekly_df,
    }
)

st.download_button(
    "Download Excel workbook",
    data=excel_payload,
    file_name="search_mix_optimizer_outputs.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
